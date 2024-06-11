'''
Why I need this python sciript?
ans:- We were exporting the logs from an application, The logs are in the format of .xlsx extensions. This files are to be renamed on the value present in the excel file. 
so, thought of creating the script for that.

What this script does?
Ans:- It opens all the files present in the folder which having extensions .xlsx. Enters in each file read the particular value, store it and then rename it with the same name. 

Modules Used here are OS, glob and Openpyxl 
'''
import os 
import glob
from openpyxl import load_workbook

f = glob.glob('*.xlsx')

for i in f:
    wb = load_workbook(filename=i)
    ws = wb.active
    new_name = ws.cell(row=2,column=4).value
    file_name = new_name + "custome_value.xlsx"

    wb.close()
    os.rename (i,file_name)

    

